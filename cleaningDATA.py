import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def browse_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel or CSV Files", "*.xlsx;*.xls;*.csv")])
    if file_path:
        file_label.config(text=f"Selected File: {file_path}")
        clean_data()

def clean_data():
    try:
        # Read file as text
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=None, dtype=str)
        else:
            df = pd.read_excel(file_path, header=None, dtype=str)

        # Remove first 20 rows
        #df = df.iloc[20:].reset_index(drop=True)
        
        # Keep only rows where first column contains exactly 8 "!"
        df = df[df[0].str.count("!") == 8]
        df = df.reset_index(drop=True)


        # Split first column by "!"
        df = df[0].str.split("!", expand=True)

        # Rename columns
        df.columns = [
            "Valeur", "Capitaux Debit", "Capitaux Credit", "Soldes Debit", "Soldes Credit", 
            "NBJ", "Nombres Debit", "Nombres Credit", "Taux Decouvert"
        ]

        # Clean date function
        def clean_date(value):
            if isinstance(value, str):
                value = value.replace('*', '').strip()
                try:
                    return datetime.strptime(value, "%d/%m/%y")
                except ValueError:
                    return None  # Keep invalid dates as NaN
            return value

        # Apply date cleaning
        df["Valeur"] = df["Valeur"].apply(clean_date)

        # Fill missing dates in "Valeur" column
        df["Valeur"] = df["Valeur"].ffill()

        # Convert numeric columns
        def convert_to_number(value):
            if isinstance(value, str):
                value = value.replace('.', '').replace(',', '.').strip()
            try:
                return float(value)
            except ValueError:
                return None  # Keep invalid numbers as NaN

        for col in df.columns[1:]:  # Skip first column (date)
            df[col] = df[col].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
            df[col] = df[col].apply(convert_to_number)
            
            
        # Divide numeric values by 100
        for col in df.columns[1:-1]:
           if col != 'NBJ':
              df[col] = df[col] / 100

        # Divide by 10,000,000 for the last column (Taux Decouvert)
        df["Taux Decouvert"] = df["Taux Decouvert"] / 10000000


        # Remove completely empty rows
        df.dropna(how='all', inplace=True)

        # Remove rows that contain only a date but no other values
        df = df[~((df.iloc[:, 0].notna()) & (df.iloc[:, 1:].isna().all(axis=1)))]
        print(type(df))  # Vérifie si df est bien un DataFrame


        # Save cleaned file
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        dir_path = os.path.dirname(file_path)
        cleaned_file_path = os.path.join(dir_path, f"cleaned_data_{timestamp}.xlsx")

        with pd.ExcelWriter(cleaned_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, startrow=2)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

        # Load workbook for formatting
        wb = load_workbook(cleaned_file_path)
        ws = wb.active

        # Set headers
        ws["A1"] = "Valeur"
        ws["B1"] = "Capitaux"
        ws["D1"] = "Soldes"
        ws["F1"] = "NBJ"
        ws["G1"] = "Nombres"
        ws["I1"] = "Taux Decouvert"

        # Merge headers
        ws.merge_cells("B1:C1")  # Capitaux
        ws.merge_cells("D1:E1")  # Soldes
        ws.merge_cells("G1:H1")  # Nombres

        # Set sub-headers
        sub_headers = ["Date", "Debit", "Credit", "Debit", "Credit", "NBJ", "Debit", "Credit", "Taux Decouvert"]
        for col_num, sub_header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col_num, value=sub_header)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 
        ws.column_dimensions['A'].width = 12
        
        # Apply number format to columns B to I
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            for cell in ws[col_letter]:
                if isinstance(cell.value, (int, float)):
                   cell.number_format = '#,##0.00'

        
        

        
    
        
        for cell in ws["A"]:
          if isinstance(cell.value, datetime):
                cell.number_format = "DD/MM/YYYY"

   
        
       


        # Save workbook
        wb.save(cleaned_file_path)
        messagebox.showinfo("Success", f"Data cleaning completed!\nSaved at:\n{cleaned_file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# GUI Setup
window = tk.Tk()
window.title("Excel File Selector")
window.geometry("800x400")

browse_button = tk.Button(window, text="Browse Excel or CSV File", command=browse_file)
browse_button.pack(pady=10)

file_label = tk.Label(window, text="No file selected")
file_label.pack()
window.mainloop()